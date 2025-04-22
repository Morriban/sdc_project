import os
import json
from openpyxl import Workbook, load_workbook
from encryption_module import encrypt_data
from acm_module import generate_acm

# These global variables are used to determine the file directory
DATA_DIR = "data"
DEFAULT_SDC_PATH = os.path.join(DATA_DIR, "SecureDataContainer.xlsx")


# This helper function is used to encrypt and embed the ACM
def embed_acm_in_workbook(wb: Workbook, acm: dict, acm_key: bytes):
    acm_ws = wb.create_sheet("_ACM")
    acm_ws.sheet_state = "hidden"

    acm_json = json.dumps(acm)
    encrypted_acm = encrypt_data(acm_key, acm_json)
    acm_ws["A1"] = encrypted_acm


# This function creates a new workbook and fills it with 5 sheets and some data
# It then goes through each sheet, creating a new key, and checks each cell for data.
# If the cell has data, then it will cal encrypt_data to encrypt it.
# It will then create a key for the ACM and embed that into the sheet.
# Finally, it will create a json file to save the keys and save the new SDC.
def create_sdc(output_file=DEFAULT_SDC_PATH):
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = Workbook()
    sheet_names = ["Sheet1", "Sheet2", "Sheet3", "Sheet4", "Sheet5"]
    ws = wb.active
    ws.title = sheet_names[0]

    for name in sheet_names[1:]:
        wb.create_sheet(title=name)

    for sheet in wb.sheetnames:
        if sheet != "_ACM":
            ws = wb[sheet]
            for row in range(1, 6):
                ws[f"A{row}"] = f"This is {sheet} - Row {row}"

    keys = {}
    for sheet in sheet_names:
        ws = wb[sheet]
        key = os.urandom(32)
        keys[sheet] = key.hex()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = encrypt_data(key, str(cell.value))

    acm_key = os.urandom(32)
    acm = generate_acm(sheet_names)
    embed_acm_in_workbook(wb, acm, acm_key)

    key_file = output_file.replace(".xlsx", "_keys.json")
    keys["__acm__"] = acm_key.hex()
    with open(key_file, "w") as f:
        json.dump(keys, f)

    wb.save(output_file)
    print(f"Blank SDC created at: {output_file}")


# This function creates a new workbook filled with the same contents of the inputted file
# It then goes through each sheet, creating a new key, and checks each cell for data.
# If the cell has data, then it will cal encrypt_data to encrypt it.
# It will then create a key for the ACM and embed that into the sheet.
# Finally, it will create a json file to save the keys and save the new SDC.
def encrypt_existing_excel(file_path, sdc_name="default_sdc"):
    sdc_dir = f"data/sdcs/{sdc_name}"
    os.makedirs(sdc_dir, exist_ok=True)

    output_file = os.path.join(sdc_dir, "SecureDataContainer.xlsx")
    wb = load_workbook(file_path)
    keys = {}

    for sheet in wb.sheetnames:
        if sheet.startswith("_"):
            continue
        ws = wb[sheet]
        key = os.urandom(32)
        keys[sheet] = key.hex()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.value = encrypt_data(key, str(cell.value))

    acm_key = os.urandom(32)
    acm = generate_acm([s for s in wb.sheetnames if not s.startswith("_")])
    embed_acm_in_workbook(wb, acm, acm_key)

    key_file = output_file.replace(".xlsx", "_keys.json")
    keys["__acm__"] = acm_key.hex()
    with open(key_file, "w") as f:
        json.dump(keys, f)

    wb.save(output_file)
    print(f"Encrypted SDC saved to: {output_file}")
