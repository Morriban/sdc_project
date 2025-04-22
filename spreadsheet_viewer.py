import json
from openpyxl import load_workbook, Workbook
from encryption_module import decrypt_data


# This helper function is used to decrypt the embedded ACM
def load_acm_from_workbook(wb, acm_key: bytes):
    acm_ws = wb["_ACM"]
    encrypted_acm = acm_ws["A1"].value
    if not encrypted_acm:
        raise ValueError("Missing ACM data in workbook.")
    acm_json = decrypt_data(acm_key, encrypted_acm)
    return json.loads(acm_json)


# This helper function is used to quickly return the sheets that would be accessible to a role
def get_accessible_sheets(acm, role):
    return acm.get(role, [])


# This function takes in an Excel file and a users' role.
# It then opens the key json file and decrypts the ACM and uses it to determine what sheets to copy into a new workbook.
# It then decrypts each sheet in the workbook as only the appropriate sheets were added.
# Finally, it saves the decrypted SDC.
def view_sdc(role, input_file):
    output_file = input_file.replace(".xlsx", f"_decrypted_{role}.xlsx")

    # Load external key file
    key_file = input_file.replace(".xlsx", "_keys.json")
    with open(key_file, "r") as f:
        keys = json.load(f)

    acm_key = bytes.fromhex(keys["__acm__"])
    wb = load_workbook(input_file)
    acm = load_acm_from_workbook(wb, acm_key)
    accessible_sheets = get_accessible_sheets(acm, role)

    decrypted_wb = Workbook()
    decrypted_wb.remove(decrypted_wb.active)

    for sheet in accessible_sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        new_ws = decrypted_wb.create_sheet(sheet)

        key = bytes.fromhex(keys.get(sheet))
        for row in ws.iter_rows():
            new_row = []
            for cell in row:
                if cell.value:
                    try:
                        decrypted_value = decrypt_data(key, cell.value)
                        new_row.append(decrypted_value)
                    except:
                        new_row.append("[DECRYPT_ERROR]")
                else:
                    new_row.append("")
            new_ws.append(new_row)

    decrypted_wb.save(output_file)
    print(f"Decrypted SDC for role '{role}' saved to: {output_file}")
